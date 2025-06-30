# -*- coding: utf-8 -*-
"""
@Author: D
@Date: 2025/6/12
@Describe
"""
import os
import io
import requests

import cv2
import torch
import numpy as np
from PIL import Image, ImageSequence, ImageOps
from openpyxl import Workbook
from openpyxl import load_workbook

import node_helpers
import folder_paths


class LoadImageWithAlpha:
    @classmethod
    def INPUT_TYPES(cls):
        input_dir = folder_paths.get_input_directory()
        files = [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))]
        files = folder_paths.filter_files_content_types(files, ["image"])
        return {
            "required": {
                "image": (sorted(files), {"image_upload": True}),
                "retain_alpha": ("BOOLEAN", {"default": True}),
            }
        }



    RETURN_TYPES = ("IMAGE", "MASK")
    RETURN_NAMES = ("image", "mask")
    FUNCTION = "load_image_with_alpha"
    CATEGORY = "图像处理☕️"
    DESCRIPTION = """读取图片，保留alpha通道"""

    def load_image_with_alpha(self, image, retain_alpha):
        image_path = folder_paths.get_annotated_filepath(image)

        img = node_helpers.pillow(Image.open, image_path)

        output_images = []
        output_masks = []
        w, h = None, None

        excluded_formats = ['MPO']
        for i in ImageSequence.Iterator(img):
            i = node_helpers.pillow(ImageOps.exif_transpose, i)

            if i.mode == 'I':
                i = i.point(lambda i: i * (1 / 255))

            image = i.convert("RGB")

            has_alpha = "A" in i.getbands()
            if retain_alpha:
                if has_alpha:
                    image = i.convert("RGBA")
                if i.mode == 'P' and 'transparency' in i.info:
                    image = i.convert("RGBA")

            if len(output_images) == 0:
                w = image.size[0]
                h = image.size[1]

            if image.size[0] != w or image.size[1] != h:
                continue

            image = np.array(image).astype(np.float32) / 255.0
            image = torch.from_numpy(image)[None,]
            if 'A' in i.getbands():
                mask = np.array(i.getchannel('A')).astype(np.float32) / 255.0
                mask = 1. - torch.from_numpy(mask)
            elif i.mode == 'P' and 'transparency' in i.info:
                mask = np.array(i.convert('RGBA').getchannel('A')).astype(np.float32) / 255.0
                mask = 1. - torch.from_numpy(mask)
            else:
                mask = torch.zeros((64, 64), dtype=torch.float32, device="cpu")
            output_images.append(image)
            output_masks.append(mask.unsqueeze(0))

        if len(output_images) > 1 and img.format not in excluded_formats:
            output_image = torch.cat(output_images, dim=0)
            output_mask = torch.cat(output_masks, dim=0)
        else:
            output_image = output_images[0]
            output_mask = output_masks[0]

        return (output_image, output_mask)


class ImageResizeMaxLen:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image": ("IMAGE",),
                "max_len": ("INT", {"default": 640, "min": 32, "max": 5000, "step": 1}),
            }
        }

    RETURN_TYPES = ("IMAGE",)
    FUNCTION = "image_resize"
    CATEGORY = "图像处理☕️"
    DESCRIPTION = """将图像按最长边调整到指定大小，保持原始长宽比, 最小32最大5000"""

    def image_resize(self, image, max_len):
        # 确保输入是 NumPy 数组并且具有正确的形状
        if len(image.shape) == 4:
            image = image.squeeze(0)  # 移除批次维度如果存在

        # 将图像转换为 NumPy 数组如果它还不是
        if isinstance(image, torch.Tensor):
            image = image.cpu().numpy()

        # image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
        (h, w) = image.shape[:2]
        if w >= h:
            ratio = max_len / float(w)
            dim = (max_len, int(h * ratio))
        else:
            ratio = max_len / float(h)
            dim = (int(w * ratio), max_len)

        image_resize = cv2.resize(image, dim, interpolation=cv2.INTER_CUBIC)
        image_resize_tensor = torch.from_numpy(image_resize).unsqueeze(0).float()

        return (image_resize_tensor,)


class ImageResizeWidthHeight:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image": ("IMAGE",),
                "width": ("INT", {"default": 640, "min": 32, "max": 5000, "step": 1}),
                "retain_width": ("BOOLEAN", {"default": False}),
                "height": ("INT", {"default": 640, "min": 32, "max": 5000, "step": 1}),
                "retain_height": ("BOOLEAN", {"default": False}),
            }
        }

    RETURN_TYPES = ("IMAGE",)
    FUNCTION = "image_resize"
    CATEGORY = "图像处理☕️"
    DESCRIPTION = """将图像按指定长宽调整大小，会破坏图片长宽比"""

    def image_resize(self, image, width, height, retain_width, retain_height):
        # 确保输入是 NumPy 数组并且具有正确的形状
        if len(image.shape) == 4:
            image = image.squeeze(0)  # 移除批次维度如果存在

        # 将图像转换为 NumPy 数组如果它还不是
        if isinstance(image, torch.Tensor):
            image = image.cpu().numpy()

        # image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
        (h, w) = image.shape[:2]
        if retain_width:
            dim = (w, height)
        elif retain_height:
            dim = (width, h)
        else:
            dim = (width, height)

        image_resize = cv2.resize(image, dim, interpolation=cv2.INTER_CUBIC)
        image_resize_tensor = torch.from_numpy(image_resize).unsqueeze(0).float()

        return (image_resize_tensor,)


class ObtainImageSize:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image": ("IMAGE",),
            }
        }

    RETURN_TYPES = ("INT", "INT", "BOOLEAN")
    RETURN_NAMES = ("width", "height","is_alpha")
    FUNCTION = "obtain_image_size"
    CATEGORY = "图像处理☕️"
    DESCRIPTION = """获取图片的宽高，且判断是否包含透明通道"""

    def obtain_image_size(self, image):
        # 确保输入是 NumPy 数组并且具有正确的形状
        if len(image.shape) == 4:
            image = image.squeeze(0)  # 移除批次维度如果存在

        # 将图像转换为 NumPy 数组如果它还不是
        if isinstance(image, torch.Tensor):
            image = image.cpu().numpy() * 255

        (h, w) = image.shape[:2]
        image_pil = Image.fromarray(image.astype(np.uint8)).convert("RGBA")
        a = image_pil.getchannel('A')
        if any(pixel == 0 for pixel in a.getdata()):
            is_alpha = True
        else:
            is_alpha = False

        return (w, h, is_alpha)


class ObtainImageSizeFromXls:
    def __init__(self):
        self.output_dir = folder_paths.get_output_directory()

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "path": ("STRING",),
                "save_name": ("STRING", {"default": 'output.xlsx'})
            }
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("recall",)
    FUNCTION = "obtain_image_size_xls"
    CATEGORY = "图像处理☕️"
    DESCRIPTION = """从xls文件获取图片，输出图片的宽高，且判断是否包含透明通道，写入新的xls文件"""

    def obtain_image_size_xls(self, path, save_name):
        # 检查保存文件名
        if save_name[-4:] != '.xls' and save_name[-5:] != '.xlsx':
            save_name = save_name + '.xlsx'
        save = os.path.join(self.output_dir, save_name)

        # 读取xls文件
        rs = self.read_xls(path, 10000)

        d = []
        for r in rs:
            img_url = r[0]
            im_array, alp = self.download_image(img_url)
            if im_array is None:
                continue
            whc = self.get_size(im_array, alp)
            result = [img_url] + whc
            d.append(result)

        self.write_xls(d, save)
        recall = "完成任务"

        return (recall,)

    def read_xls(self, file_path, max_r):
        result = []
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(max_col=100, max_row=max_r, values_only=True):
            result.append(row)
        return result

    def download_image(self, url):
        f = False
        ee = ''
        tt = 10
        while not f:
            try:
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
                    "Referer": "https://www.threadless.com/"
                }
                response = requests.get(url, headers=headers, stream=True)
                response.raise_for_status()

                image_bytes = io.BytesIO(response.content)
                image = Image.open(image_bytes).convert('RGBA')
                a = image.getchannel('A')
                rgb_array = np.array(image)
                f = True
                return rgb_array, a
            except Exception as e:
                ee = e
                if tt == 0:
                    break
                tt -= 1
        print(f"下载图片失败 {url}: {str(ee)}")
        return None, None

    def get_size(self, img_array, alpha):
        h, w, c = img_array.shape
        if any(pixel == 0 for pixel in alpha.getdata()):
            is_alpha = 'yes'
        else:
            is_alpha = 'no'

        return [h, w, is_alpha]

    def write_xls(self, data, file):
        wb = Workbook()
        ws = wb.active

        # 批量写入数据
        for i, row in enumerate(data):
            ws.append(row)

        # 保存文件
        wb.save(file)


class CropAlphaImage:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image": ("IMAGE",),
                "pixels": ("INT", {"default": 10, "min": 0, "step": 1, }),
                "center": ("BOOLEAN", {"default": True})
            }
        }

    RETURN_TYPES = ("IMAGE", )
    RETURN_NAMES = ("crop_image", )
    FUNCTION = "crop_alpha_image"
    CATEGORY = "图像处理☕️"
    DESCRIPTION = """按指定像素边距裁剪图片透明通道"""

    def crop_alpha_image(self, image, pixels, center):
        # 确保输入是 NumPy 数组并且具有正确的形状
        print(image.shape)
        if len(image.shape) == 4:
            image = image.squeeze(0)  # 移除批次维度如果存在

        # 将图像转换为 NumPy 数组如果它还不是
        if isinstance(image, torch.Tensor):
            image = image.cpu().numpy()

        h, w, c = image.shape
        # image_pil = Image.fromarray(image.astype(np.uint8)).convert("RGBA")
        # a = image_pil.getchannel('A')
        if c == 3:  # RGB2RGBA
            image_tensor = torch.from_numpy(image).unsqueeze(0)
        else:
            rgba_array = image

            # 获取非透明像素边界
            xmin, ymin, xmax, ymax = None, None, None, None
            for x in range(w):
                for y in range(h):
                    if rgba_array[y, x, 3] != 0:
                        if xmin is None:
                            xmin = x
                        if ymin is None:
                            ymin = y
                        if xmax is None:
                            xmax = x
                        if ymax is None:
                            ymax = y

                        if xmin > x:
                            xmin = x
                        if xmax < x:
                            xmax = x
                        if ymin > y:
                            ymin = y
                        if ymax < y:
                            ymax = y

            # 计算裁剪边界
            c_w, c_h = pixels, pixels
            if center:  # 居中裁剪
                if round(xmin - c_w) < 0:
                    c_w = xmin
                if round(xmax + c_w) > w:
                    c_w = w - xmax
                if round(ymin - c_h) < 0:
                    c_h = ymin
                if round(ymax + c_h) > h:
                    c_h = h - ymax
                xmin = round(xmin - c_w)
                ymin = round(ymin - c_h)
                xmax = round(xmax + c_w)
                ymax = round(ymax + c_h)
            else:  # 严格按照指定边距裁剪，可能不居中
                xmin = max(round(xmin - c_w), 0)
                ymin = max(round(ymin - c_h), 0)
                xmax = min(round(xmax + c_w), w)
                ymax = min(round(ymax + c_h), h)

            crop_image = rgba_array[ymin: ymax, xmin: xmax, :]
            image_tensor = torch.from_numpy(crop_image).unsqueeze(0).float()

        return (image_tensor,)


IMAGE_NODE_CLASS_MAPPINGS = {
    "保留alpha通道加载图像": LoadImageWithAlpha,
    "指定最长边调整图像大小": ImageResizeMaxLen,
    "指定长宽调整图像大小": ImageResizeWidthHeight,
    "获取图像尺寸透明通道": ObtainImageSize,
    "获取图像尺寸透明通道xls": ObtainImageSizeFromXls,
    "裁剪图片透明通道": CropAlphaImage,
}

IMAGE_NODE_DISPLAY_NAME_MAPPINGS = {
    "保留alpha通道加载图像": "保留alpha通道加载图像",
    "指定最长边调整图像大小": "指定最长边调整图像大小",
    "指定长宽调整图像大小": "指定长宽调整图像大小",
    "获取图像尺寸透明通道": "获取图像尺寸透明通道",
    "获取图像尺寸透明通道xls": "获取图像尺寸透明通道xls",
    "裁剪图片透明通道": "裁剪图片透明通道",
}
